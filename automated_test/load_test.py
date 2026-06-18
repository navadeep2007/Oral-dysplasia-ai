import os
import sys
import time
import datetime
import urllib.request
import urllib.error
import threading
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Target URL - target the local FastAPI backend server
TARGET_URL = "http://127.0.0.1:8000/health"
VIRTUAL_USERS = 100
DURATION_SECONDS = 60
OUTPUT_FILE = os.path.join("automated_test", "Load_Test_Report.xlsx")

results = []
results_lock = threading.Lock()
stop_event = threading.Event()

def virtual_user_task(user_id):
    """Simulates a single virtual user making concurrent requests."""
    while not stop_event.is_set():
        req_start_time = datetime.datetime.now()
        start = time.perf_counter()
        status = 0
        try:
            req = urllib.request.Request(
                TARGET_URL, 
                headers={'User-Agent': f'LoadTester/1.0 User-{user_id}'}
            )
            with urllib.request.urlopen(req, timeout=5) as response:
                status = response.status
                response.read()
        except urllib.error.HTTPError as e:
            status = e.code
        except urllib.error.URLError as e:
            status = 503  # Service unavailable
        except Exception:
            status = 500  # Connection errors, etc.
        
        end = time.perf_counter()
        latency = (end - start) * 1000.0  # Convert to ms
        
        with results_lock:
            results.append({
                "timestamp": req_start_time,
                "user_id": user_id,
                "status": status,
                "latency": latency
            })
        
        # 10ms sleep to control the burst and simulate realistic behavior
        time.sleep(0.01)

def format_excel_sheet(ws):
    """Enable gridlines and apply auto-fit column widths."""
    ws.views.sheetView[0].showGridLines = True
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def generate_excel_report(duration_actual):
    """Processes the test results and writes a formatted Excel report."""
    total_requests = len(results)
    if total_requests == 0:
        print("[WARN] No requests were sent during the test.")
        return

    successes = [r for r in results if r["status"] == 200]
    total_success = len(successes)
    total_failed = total_requests - total_success
    success_rate = (total_success / total_requests) if total_requests > 0 else 0
    
    latencies = [r["latency"] for r in results]
    avg_latency = sum(latencies) / total_requests if total_requests > 0 else 0
    min_latency = min(latencies) if latencies else 0
    max_latency = max(latencies) if latencies else 0
    
    rps = total_requests / duration_actual if duration_actual > 0 else 0

    wb = Workbook()
    
    # ── SHEET 1: DASHBOARD SUMMARY ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary Dashboard"
    
    # Colors
    navy_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    # Fonts
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    double_bottom_border = Border(
        bottom=Side(style='double', color='000000'),
        top=Side(style='thin', color='BFBFBF')
    )

    # Title Banner
    ws1.merge_cells("A1:G2")
    title_cell = ws1["A1"]
    title_cell.value = "OralDysplasia AI — Baseline Load Test Report"
    title_cell.font = title_font
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply fill to other merged cells in title banner
    for r in range(1, 3):
        for c in range(1, 8):
            ws1.cell(row=r, column=c).fill = navy_fill

    # Test Metadata (Row 4-8)
    metadata = [
        ("Target URL", TARGET_URL),
        ("Virtual Users", VIRTUAL_USERS),
        ("Test Target Duration", f"{DURATION_SECONDS} seconds"),
        ("Actual Test Duration", f"{duration_actual:.2f} seconds"),
        ("Execution Date & Time", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ]
    for idx, (label, val) in enumerate(metadata, start=4):
        ws1.cell(row=idx, column=1, value=label).font = bold_font
        ws1.cell(row=idx, column=1).alignment = Alignment(horizontal="left")
        ws1.cell(row=idx, column=2, value=val).font = normal_font
        ws1.cell(row=idx, column=2).alignment = Alignment(horizontal="left")

    # Section Header: Performance Metrics
    ws1.cell(row=10, column=1, value="Core Performance Metrics").font = section_font
    
    # KPI Headers
    kpis = [
        ("Total Requests", total_requests, "#,##0"),
        ("Successful Requests (200)", total_success, "#,##0"),
        ("Failed Requests", total_failed, "#,##0"),
        ("Success Rate", success_rate, "0.0%"),
        ("Average Response Time", avg_latency, "0.00\" ms\""),
        ("Min Response Time", min_latency, "0.00\" ms\""),
        ("Max Response Time", max_latency, "0.00\" ms\""),
        ("Requests Per Second (RPS)", rps, "0.00")
    ]
    
    # Create metric table block
    ws1.cell(row=11, column=1, value="Metric").font = header_font
    ws1.cell(row=11, column=1).fill = navy_fill
    ws1.cell(row=11, column=1).alignment = Alignment(horizontal="left")
    ws1.cell(row=11, column=1).border = thin_border
    
    ws1.cell(row=11, column=2, value="Value").font = header_font
    ws1.cell(row=11, column=2).fill = navy_fill
    ws1.cell(row=11, column=2).alignment = Alignment(horizontal="right")
    ws1.cell(row=11, column=2).border = thin_border
    
    for idx, (metric, val, fmt) in enumerate(kpis, start=12):
        c1 = ws1.cell(row=idx, column=1, value=metric)
        c2 = ws1.cell(row=idx, column=2, value=val)
        
        c1.font = bold_font if "Rate" in metric or "RPS" in metric or "Total" in metric else normal_font
        c2.font = bold_font if "Rate" in metric or "RPS" in metric or "Total" in metric else normal_font
        
        c1.border = thin_border
        c2.border = thin_border
        c2.number_format = fmt
        
        c1.alignment = Alignment(horizontal="left")
        c2.alignment = Alignment(horizontal="right")
        
        # Color highlighting
        if "Success Rate" in metric:
            fill = green_fill if val >= 0.99 else red_fill
            c1.fill = fill
            c2.fill = fill
        elif "Failed" in metric and val > 0:
            c1.fill = red_fill
            c2.fill = red_fill

    # Status Code Distribution Table
    ws1.cell(row=22, column=1, value="Status Code Distribution").font = section_font
    
    ws1.cell(row=23, column=1, value="Status Code").font = header_font
    ws1.cell(row=23, column=1).fill = navy_fill
    ws1.cell(row=23, column=1).border = thin_border
    
    ws1.cell(row=23, column=2, value="Count").font = header_font
    ws1.cell(row=23, column=2).fill = navy_fill
    ws1.cell(row=23, column=2).border = thin_border
    
    ws1.cell(row=23, column=3, value="Percentage").font = header_font
    ws1.cell(row=23, column=3).fill = navy_fill
    ws1.cell(row=23, column=3).border = thin_border
    
    # Calculate status distribution
    status_counts = {}
    for r in results:
        status_counts[r["status"]] = status_counts.get(r["status"], 0) + 1
        
    sorted_status = sorted(status_counts.items())
    current_row = 24
    for status, count in sorted_status:
        c1 = ws1.cell(row=current_row, column=1, value=status)
        c2 = ws1.cell(row=current_row, column=2, value=count)
        c3 = ws1.cell(row=current_row, column=3, value=count / total_requests)
        
        c1.font = normal_font
        c2.font = normal_font
        c3.font = normal_font
        
        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border
        
        c1.alignment = Alignment(horizontal="center")
        c2.alignment = Alignment(horizontal="right")
        c3.alignment = Alignment(horizontal="right")
        
        c2.number_format = "#,##0"
        c3.number_format = "0.0%"
        
        if status == 200:
            c1.fill = green_fill
            c2.fill = green_fill
            c3.fill = green_fill
        else:
            c1.fill = red_fill
            c2.fill = red_fill
            c3.fill = red_fill
            
        current_row += 1
        
    format_excel_sheet(ws1)

    # ── SHEET 2: RAW DATA ───────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Raw Latency Data")
    
    headers = ["Request ID", "Timestamp", "Virtual User ID", "Response Status", "Latency (ms)"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    for idx, r in enumerate(results, start=2):
        c1 = ws2.cell(row=idx, column=1, value=idx - 1)
        c2 = ws2.cell(row=idx, column=2, value=r["timestamp"].strftime("%Y-%m-%d %H:%M:%S.%f")[:-3])
        c3 = ws2.cell(row=idx, column=3, value=r["user_id"])
        c4 = ws2.cell(row=idx, column=4, value=r["status"])
        c5 = ws2.cell(row=idx, column=5, value=r["latency"])
        
        for cell in [c1, c2, c3, c4, c5]:
            cell.font = normal_font
            cell.border = thin_border
            
        c1.alignment = Alignment(horizontal="center")
        c2.alignment = Alignment(horizontal="center")
        c3.alignment = Alignment(horizontal="center")
        c4.alignment = Alignment(horizontal="center")
        c5.alignment = Alignment(horizontal="right")
        
        c5.number_format = "0.0"
        
        # Color highlight bad responses
        if r["status"] != 200:
            c4.fill = red_fill
            
    format_excel_sheet(ws2)
    
    # Save the workbook
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"[OK] Excel report generated successfully at: {OUTPUT_FILE}")

def main():
    start_time = time.time()
    
    threads = []
    # Create thread workers for simulating concurrent virtual users
    for i in range(1, VIRTUAL_USERS + 1):
        t = threading.Thread(target=virtual_user_task, args=(i,))
        threads.append(t)
        t.start()
        
    print(f"Running concurrent load test with {VIRTUAL_USERS} virtual users for 60 seconds...")
    try:
        # Wait for duration
        time.sleep(DURATION_SECONDS)
    except KeyboardInterrupt:
        print("\n[WARN] Test interrupted by user.")
    finally:
        stop_event.set()
        
    for t in threads:
        t.join()
        
    actual_duration = time.time() - start_time
    
    total_req = len(results)
    successes = [r for r in results if r["status"] == 200]
    failed = total_req - len(successes)
    avg_lat = sum(r["latency"] for r in results) / total_req if total_req > 0 else 0
    min_lat = min(r["latency"] for r in results) if results else 0
    max_lat = max(r["latency"] for r in results) if results else 0
    rps_rate = total_req / actual_duration if actual_duration > 0 else 0
    
    print("\n" + "=" * 40)
    print("        LOAD TEST RESULTS SUMMARY")
    print("=" * 40)
    print(f"[OK] Target URL:             {TARGET_URL}")
    print(f"[OK] Virtual Users:          {VIRTUAL_USERS}")
    print(f"[OK] Actual Test Duration:   {actual_duration:.2f} seconds")
    print(f"[OK] Total Requests Sent:    {total_req}")
    print(f"[OK] Successful Requests:    {len(successes)}")
    print(f"[FAIL] Failed Requests:      {failed}")
    if total_req > 0:
        print(f"[OK] Success Rate:           {(len(successes)/total_req)*100:.2f}%")
        print(f"[OK] Requests Per Second:    {rps_rate:.2f} RPS")
        print(f"[OK] Avg Response Time:      {avg_lat:.2f} ms")
        print(f"[OK] Min Response Time:      {min_lat:.2f} ms")
        print(f"[OK] Max Response Time:      {max_lat:.2f} ms")
    else:
        print("[WARN] No requests completed successfully.")
    print("=" * 40)
    
    generate_excel_report(actual_duration)

if __name__ == "__main__":
    main()
